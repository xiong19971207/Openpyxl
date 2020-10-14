from openpyxl import load_workbook

wb = load_workbook('销售经营表.xlsx')

ws = wb['在途库存']

print(ws['A1'].value)

cell_range = ws['A1':'C2']
print(cell_range)

for rows in cell_range:
    print()
    for i in rows:
        print(i.value, end=' ')

# 找出一列值
col_c = ws['C']
print()
print(col_c)

# 切片方式取出列
col_range = ws['C':'D']
print(col_range)

# 切片取出每一行的值
row = ws[1:10]
print(row)

# 按行取出值,同时规定取出多少列
print('====================')
cells = ws.iter_cols(max_col=3, min_col=1, max_row=2)
for cell in cells:
    for row in cell:
        print(row.value)

# rows = ws.iter_rows(max_col=3,max_row=)
