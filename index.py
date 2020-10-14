# 测试openpyxl
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

# 为excel的sheet添加颜色
ws.sheet_properties.tabColor = '00ff00'

ws['A1'] = 10086

print(wb.sheetnames)

wb.copy_worksheet(wb.active)

# 获取某一个单元格的值
c = ws['A1'].value
print(c)

# 通过行和列来获取某个单元格的值
c = ws.cell(column=1, row=1,).value
print(c)

wb.save('test.xlsx')
